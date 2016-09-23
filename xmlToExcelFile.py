from xml.etree import ElementTree
from openpyxl import Workbook
from easygui import *
import sys

# Initializing the Excel workbook
wb = Workbook()
# Pointing to the file that XML needs to be parsed and extracted to
writeToFile = 'Unlocked.xlsx'

# Prompting the user for the locaion to the XML file on disk and parsing it
filePath = enterbox("Enter the path to XML file")
xmlFile = open(filePath)
studentRecords = ElementTree.parse(xmlFile)

# Finding root element of XML
root = studentRecords.getroot()

# Capturing the list of all the StudentRecords
for student in root.findall('StudentRecord'):

	# Locating and capturing the CSSIDNumber from the list of children
	cssid = student.find('Student').find('PersonalIdentifiers').findtext('CSSIDNumber')
	ws = wb.create_sheet(str(cssid), 0) # Assigning CSSIDNumber as the sheet name
	ws.cell(row = 1, column = 1, value = "Tag") # Providing column name
	ws.cell(row = 1, column = 2, value = "Value") # Providing column name
	rowNumber = 2
	columnNumber = 1

	# Iterating through and capturing information about every child of each StudentRecord tag
	for record in student.iter():
		ws.cell(row = rowNumber, column = columnNumber, value = record.tag) # Printing tag name
		columnNumber = columnNumber + 1
		ws.cell(row = rowNumber, column = columnNumber, value = record.text) # Printing text between tags
		rowNumber = rowNumber + 1
		columnNumber = columnNumber - 1
wb.save(writeToFile)